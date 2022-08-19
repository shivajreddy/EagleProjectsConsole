# COMPARING ALGORITHM - MAIN FILE_V6 - ARCHITECTONIC - </> by ShivaReddy.
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
# import pillow

#####
# PART 1 - (START) - GET THE DATA, & CONVERT IT INTO DICTIONARIES
# file_1 = pd.ExcelFile(r'C:\Users\sreddy\Desktop\testing\TestV1.xlsm')
# file_2 = pd.ExcelFile(r'C:\Users\sreddy\Desktop\testing\TestV2.xlsm')
# r_file_path = r'C:\Users\sreddy\Desktop\t2.xlsm'
# TODO Use these paths for test. Later these are sent from the run algo view
# file_1_name = "C:\Users\sreddy\Desktop\EagleProjectsConsole\app\static\uploaded_aqt_files\Lot-014-03-RB_CDR-20_V03_AQT_2022-08-111660237536102.xlsm"
# file_2_name = "C:\Users\sreddy\Desktop\EagleProjectsConsole\app\static\uploaded_aqt_files\Lot-014-03-RB_CDR-20_V03_AQT_2022-08-121660237536102.xlsm"
# result_path = "C:\Users\sreddy\Desktop\EagleProjectsConsole\app\static\generated_reports"

#! original paths
# file_1 = pd.ExcelFile(r'C:\Users\sreddy\OneDrive - Eagle Construction of VA\Automation Estimating Local\newtest.xlsm')
# file_2 = pd.ExcelFile(r'C:\Users\sreddy\OneDrive - Eagle Construction of VA\Automation Estimating Local\newtest2.xlsm')
# r_file_path = r'C:\Users\sreddy\OneDrive - Eagle Construction of VA\Automation Estimating Local\compare_test4.xlsm'

# file_1 = pd.ExcelFile(r"C:\Users\sreddy\Desktop\EagleProjectsConsole\app\static\uploaded_aqt_files\Main_EST_Lot-003-04-PV_CVX-10_AQT_2022-08-18.xlsm")
# file_2 = pd.ExcelFile(r"C:\Users\sreddy\Desktop\EagleProjectsConsole\app\static\uploaded_aqt_files\Main_EST_Lot-003-04-PV_CVX-10_AQT_2022-08-19.xlsm")
r_file_path = r"C:\Users\sreddy\Desktop\EagleProjectsConsole\app\static\generated_reports\ChangeReport_Template.xlsm"


# PART 1.1 - CHOOSE THE TWO FILES
file_1_location = r''
file_2_location = r''


# PART 1.2 - GET ALL THE SHEET NAMES FROM THE EXCEL FILES
def sheet_names(location_1, location_2):
    pd_file_1 = pd.ExcelFile(location_1)
    pd_file_2 = pd.ExcelFile(location_2)
    names_of_sheets = pd_file_1.sheet_names
    # names_of_sheet2 = pd_file_2.sheet_names
    names_of_sheets.remove(names_of_sheets[0])
    return names_of_sheets

def create_data_frames(input_file_1, input_file_2, sheet_name):
    df1_1 = pd.read_excel(input_file_1, sheet_name, header=None)
    print(df1_1.shape)
    print(df1_1.shape[0])
    print(df1_1.shape[1])
    # no.of rows, columns = df1_1.shape. Skip if there is no 3rd row
    if df1_1.shape[0] == 2:
        return None
    df1_1 = df1_1.drop([0, 1, 2, ((df1_1.shape[0]) - 1)])
    df1_1 = df1_1.where(pd.notnull(df1_1), "(none)")
    f1_s1 = df1_1.to_dict(orient='records')

    df2_1 = pd.read_excel(input_file_2, sheet_name, header=None)
    if df2_1.shape[0] == 2:
        return None
    df2_1 = df2_1.drop([0, 1, 2, ((df2_1.shape[0]) - 1)])
    df2_1 = df2_1.where(pd.notnull(df2_1), "(none)")
    f2_s1 = df2_1.to_dict(orient='records')

    return f1_s1, f2_s1

excel_sheet_number = 7


def all_sheets_data_frames(input_file_1, input_file_2, excel_workbook_number):
    i = excel_workbook_number
    list_of_all_sheet_names = sheet_names(input_file_1, input_file_2)
    file1_data_frames = []
    file2_data_frames = []
    data_frames_for_2versions_single_sheet = create_data_frames(input_file_1, input_file_2, list_of_all_sheet_names[i])
    file1_data_frames.append(data_frames_for_2versions_single_sheet[0])
    file2_data_frames.append(data_frames_for_2versions_single_sheet[1])

    return file1_data_frames, file2_data_frames

# all_sheets_data_frames(file_1, file_2, excel_sheet_number)

# PART 1 - (END) - GET THE DATA, & CONVERT IT INTO DICTIONARIES
#####

#####
# PART 2 - (START) - COMPARING ALGORITHM

# PART 2.0 - CREATE INPUTS AND OUTPUTS FOR PART 2
# same_items = []
new_data_2 = []


# PART 2.1 - FUNCTIONS FOR FINDING SAME ITEMS IN TWO LISTS - LIST_A & LIST_B
# PART 2.1.1 - EVERY ITEM IN LIST_A AGAINST ONE ITEM IN LIST_B
def compare_single_item_to_a_list(list_a, item_of_b, return_same_items_list, ):
    for current_item in list_a:
        if item_of_b == current_item:
            return_same_items_list.append(current_item)
            list_a[list_a.index(current_item)] = "same_removed"
            new_data_2.append(item_of_b)
            item_of_b = "same_dict_removed"
        else:
            pass


# PART 2.1.2 - REPEATING 2.1.1 FOR EVERY ITEM IN LIST_B
def compare_lists_for_same_items(first_list, second_list, return_same_items_list):
    for current_item in second_list:
        compare_single_item_to_a_list(first_list, current_item, return_same_items_list)


# PART 2.1.3 - REMOVING REPLACED ITEMS FROM LIST
def remove_items_from_list(input_list):
    for i in input_list[:]:
        if i == "same_removed":
            input_list.remove(i)
        else:
            pass


# PART 2.1.4 - REMOVING REPLACED ITEMS FROM LIST
def remove_items_from_list_type2(input_list):
    for i in input_list[:]:
        if i == "same_dict_removed":
            input_list.remove(i)
        else:
            pass


# PART 2.1.5 - RUNNING LIST COMPARISON FOR TWO LISTS, DELETING REPLACED ITEMS IN BOTH LISTS
def compare_same_items(list_a, list_b, return_same_items_list):
    compare_lists_for_same_items(list_a, list_b, return_same_items_list)
    remove_items_from_list(list_a)
    remove_items_from_list_type2(list_b)


# PART 2.1.6 - COMPARING THE DATA SETS FOR SAME ITEMS
def remove_same_items_from_data2(input_old_data2, input_new_data2):
    temp_data2 = input_old_data2

    for x in input_new_data2[:]:
        if x in temp_data2:
            temp_data2.remove(x)

    return temp_data2


# PART 2.2.0 - FUNCTION TO COMPARE VALUES OF TWO DICTIONARIES, NOTE INDICES, GROUP INTO SUB-LISTS
def indices_of_modified(dict_a, dict_b, out_modified_indices, the_total_columns):
    for x in range(0, the_total_columns):
        if (list(dict_a.values())[x]) == (list(dict_b.values())[x]):
            out_modified_indices.append(True)
            pass
        else:
            out_modified_indices.append(False)
    return out_modified_indices


# PART 2.2.1 - EVERY DICTIONARY ITEM IN LIST_A VS. ONE DICTIONARY ITEM IN LIST_B
def compare_single_dict_to_a_list(list_a, item_of_b, return_modified_items_list, out_modified_indices, the_total_columns):
    for current_item in list_a:
        try:
            if ((list(current_item.values())[0]) == (list(item_of_b.values())[0])) & (item_of_b != current_item):
                indices_of_modified(current_item, item_of_b, out_modified_indices, the_total_columns)
                return_modified_items_list.append(item_of_b)
                list_a[list_a.index(current_item)] = "removed the dict"
                new_data_2.append(item_of_b)
                item_of_b = "removed_item_in_dict_b"
            else:
                pass
        except AttributeError:
            pass


# PART 2.2.2 - REPEATING 2.2.1 FOR EVERY DICTIONARY ITEM IN LIST_B
def compare_lists_for_modified_items(first_list, second_list, return_modified_items_list, out_modified_items, the_total_columns):
    for current_item in second_list:
        compare_single_dict_to_a_list(first_list, current_item, return_modified_items_list, out_modified_items, the_total_columns)


# PART 2.2.3 - REMOVE REPLACEMENTS IN LIST_A
def remove_dicts_from_list(input_list):
    for i in input_list[:]:
        if i == "removed the dict":
            input_list.remove(i)


# PART 2.2.4 - REMOVE REPLACEMENTS IN LIST_B
def remove_dicts_from_list_type2(input_list):
    for i in input_list[:]:
        if i == "removed_item_in_dict_b":
            input_list.remove(i)


# PART 2.2.5 - RUNNING THE COMPARISON FUNCTIONS FOR TWO DICTIONARY LISTS &  DELETING THE REPLACED ITEMS
def compare_modified_items(list_a, list_b, return_modified_items_list, out_modified_items, the_total_items):
    # tot_columns = len(list(original_data1[0].keys()))
    compare_lists_for_modified_items(list_a, list_b, return_modified_items_list, out_modified_items, the_total_items)
    remove_dicts_from_list(list_a)
    remove_dicts_from_list_type2(list_b)


# PART 2.3 - NEW ITEMS
def remove_same_and_modified_items(input_same_items, input_modified_items, original_data_2_list):
    temp_combined_list = input_same_items + input_modified_items
    temp_original_list = original_data_2_list

    for x in temp_combined_list[:]:
        if x in temp_original_list:
            temp_original_list.remove(x)

    return temp_original_list


#####
# PART 3 - (END) - SENDING DATA TO EXCEL

def comparing_algorithm(excel_file_1, excel_file_2):

    total_number_of_sheets = len(excel_file_1.sheet_names)
    total_number_of_sheets = len(excel_file_2.sheet_names)


    combined_output_same = []
    combined_output_modified = []
    combined_output_modified_indices = []
    combined_output_new = []
    combined_output_deleted = []

    for x in range(0, total_number_of_sheets-1):
        # print("doing this for sheet", x)
        data1 = all_sheets_data_frames(excel_file_1, excel_file_2, x)[0][0]
        data2 = all_sheets_data_frames(excel_file_1, excel_file_2, x)[1][0]
        original_data1 = all_sheets_data_frames(excel_file_1, excel_file_2, x)[0][0]
        original_data2 = all_sheets_data_frames(excel_file_1, excel_file_2, x)[1][0]
        number_of_columns = len(list(original_data1[0].keys()))

        same_items = []
        all_sheets_data_frames(excel_file_1, excel_file_2, x)
        compare_same_items(data1, data2, same_items)
        data2 = remove_same_items_from_data2(data2, new_data_2)

        modified_items = []
        total_columns = len(list(original_data1[0].keys()))
        modified_indices = []

        compare_modified_items(data1, data2, modified_items, modified_indices, total_columns)

        new_items = remove_same_and_modified_items(same_items, modified_items, original_data2)

        final_output_same = pd.DataFrame(same_items)
        final_output_modified = pd.DataFrame(modified_items)
        final_output_indices = modified_indices
        new_modified_indices = [modified_indices[i:i+total_columns]for i in range(0, len(modified_indices), total_columns)]
        final_output_new = pd.DataFrame(new_items)
        final_output_deleted = pd.DataFrame(data1)

        # Variables to Print Data in to Report Excel File
        wb = load_workbook(r_file_path, keep_vba= True)
        sn_list = sheet_names(excel_file_1, excel_file_2)
        ws = wb[sn_list[x]]

        combined_output_same.append(final_output_same)
        combined_output_modified.append(final_output_modified)
        combined_output_modified_indices.append(final_output_indices)
        combined_output_new.append(final_output_new)
        combined_output_deleted.append(final_output_deleted)

        # 'Same Item' Title writing
        rowindex = 3
        ws.cell(row=rowindex, column=2).value = 'Same Items'

        # 'Same Item' Title Coloring
        fill_pattern1 = PatternFill(patternType='solid', fgColor='BDD7EE')
        for i in range(len(final_output_same.columns)):
            ws.cell(row=rowindex, column=i+1).fill = fill_pattern1
        wb.save(r_file_path)

        # 'Same Item' Data writing
        for index, row in final_output_same.iterrows():
            for i in range(number_of_columns):
                ws.cell(row=rowindex+1, column=i+1).value = row.iloc[i]
            rowindex = rowindex +1
        wb.save(r_file_path)

        # 'Modified Item' Title writing
        rowindex = rowindex +2
        offset1 = rowindex
        ws.cell(row=rowindex, column=2).value = 'Modified Items'

        # 'Modified Item' Data writing
        for index, row in final_output_modified.iterrows():
            for i in range(number_of_columns):
                ws.cell(row=rowindex + 1, column=i + 1).value = row.iloc[i]
            rowindex = rowindex + 1
        wb.save(r_file_path)

        # 'Modified Item' Title Coloring
        fill_pattern1 = PatternFill(patternType='solid', fgColor='FFFF99')
        for i in range(len(final_output_same.columns)):
            ws.cell(row=offset1, column=i+1).fill = fill_pattern1
        wb.save(r_file_path)

        # 'Modified Item' modified elements coloring
        # 'Modified Item' modified elements detecting boolean mask
        indices_list = new_modified_indices
        flist = []
        for i in range(len(indices_list)):
            for j in range(len(indices_list[i])):
                if indices_list[i][j] == False:
                        flist.append([i, j])

        # print('flist: ', flist)

        # 'Modified Item' modified elements coloring by boolean mask
        fill_pattern = PatternFill(patternType='solid', fgColor='FFC000')
        for i, j in flist:
                ws.cell(row= offset1+1+i, column=j+1).fill = fill_pattern
        wb.save(r_file_path)

        # 'Added Item' Title writing
        rowindex = rowindex +2
        offset2 = rowindex
        ws.cell(row=rowindex, column=2).value = 'Added Items'

        # 'Added Item' Title Coloring
        fill_pattern1 = PatternFill(patternType='solid', fgColor='C6E0B4')
        for i in range(len(final_output_same.columns)):
            ws.cell(row=offset2, column=i+1).fill = fill_pattern1
        wb.save(r_file_path)

        # 'Added Item' data writing
        for index, row in final_output_new.iterrows():
            for i in range(number_of_columns):
                ws.cell(row=rowindex + 1, column=i + 1).value = row.iloc[i]
            rowindex = rowindex + 1
        wb.save(r_file_path)

        # 'Deleted Item' Title writing
        rowindex = rowindex +2
        offset3 = rowindex
        ws.cell(row=rowindex, column=2).value = 'Deleted Items'

        # 'Deleted Item' Title coloring
        fill_pattern1 = PatternFill(patternType='solid', fgColor='F8CBAD')
        for i in range(len(final_output_same.columns)):
            ws.cell(row=offset3, column=i+1).fill = fill_pattern1
        wb.save(r_file_path)

        # 'Deleted Item' data writing
        for index, row in final_output_deleted.iterrows():
            for i in range(number_of_columns):
                ws.cell(row=rowindex + 1, column=i + 1).value = row.iloc[i]
            rowindex = rowindex + 1
        wb.save(r_file_path)


def coverpage_data_retrieving(file_1, file_2):
    wb = load_workbook(r_file_path, keep_vba=True)
    ws = wb['00-CoverPage']
    df_cover_read = pd.read_excel(file_1, sheet_name='00-CoverPage')
    df_cover_read2 = pd.read_excel(file_2, sheet_name='00-CoverPage')
    df_cover_read=df_cover_read.fillna("")
    df_cover_read2=df_cover_read2.fillna("")

    i = 1
    rowindex = 10
    for i in range(1, 12):
          ws.cell(row=rowindex, column=4).value = df_cover_read.iloc[i,1]
          rowindex = rowindex + 1

    wb.save(r_file_path)
    #
    rowindex = 24
    for i in range(15, 23):
        ws.cell(row=rowindex, column=4).value = df_cover_read.iloc[i,1]
        rowindex = rowindex + 1
    wb.save(r_file_path)

    ws.cell(row=44, column=4).value = df_cover_read.iloc[24,1]
    ws.cell(row=45, column=4).value = df_cover_read.iloc[26,1]
    wb.save(r_file_path)

    #### Previous Version CoverPage Data to EXCEL
    i = 1
    rowindex = 10
    for i in range(1, 12):
          ws.cell(row=rowindex, column=12).value = df_cover_read2.iloc[i,1]
          rowindex = rowindex + 1

    wb.save(r_file_path)
    #
    rowindex = 24
    for i in range(15, 23):
        ws.cell(row=rowindex, column=12).value = df_cover_read2.iloc[i,1]
        rowindex = rowindex + 1
    wb.save(r_file_path)

    ws.cell(row=44, column=12).value = df_cover_read2.iloc[24,1]
    ws.cell(row=45, column=12).value = df_cover_read2.iloc[26,1]
    wb.save(r_file_path)


def three_d_model_image_importing_to_excel():
    wb = load_workbook(r_file_path, keep_vba=True)
    ws = wb['00-CoverPage']
    backimg_path = r'C:\Users\jwoo\PycharmProjects\Automation project\logos2.jpg'
    backimg_path2 = r'C:\Users\jwoo\PycharmProjects\Automation project\Changing detection\front3.png'

    backimag = openpyxl.drawing.image.Image(backimg_path)
    backimag.anchor='A1'
    ws.add_image(backimag)
    wb.save(r_file_path)

    backimag2 = openpyxl.drawing.image.Image(backimg_path2)
    backimag2.anchor='C35'
    ws.add_image(backimag2)
    wb.save(r_file_path)


# comparing_algorithm(file_1, file_2)
# coverpage_data_retrieving(file_1, file_2)
# three_d_model_image_importing_to_excel()

def run_algorithm(path1, path2):
    file_1 = pd.ExcelFile(path1)
    file_2 = pd.ExcelFile(path2)
    # file_1 = pd.ExcelFile(r"C:\Users\sreddy\Desktop\EagleProjectsConsole\app\static\uploaded_aqt_files\Main_EST_Lot-003-04-PV_CVX-10_AQT_2022-08-18.xlsm")
    # file_2 = pd.ExcelFile(r"C:\Users\sreddy\Desktop\EagleProjectsConsole\app\static\uploaded_aqt_files\Main_EST_Lot-003-04-PV_CVX-10_AQT_2022-08-19.xlsm")
    comparing_algorithm(file_1, file_2)
    coverpage_data_retrieving(file_1, file_2)
